import os
import time
import undetected_chromedriver as uc

from openpyxl import load_workbook
from bs4 import BeautifulSoup
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By


# VARIABLES
######################################################################################################################
service = Service(ChromeDriverManager().install())



chromedriver_path = "chromedriver.exe"  # Path to ChromeDriver executable
xlsx_path = "tx kansas mo .xlsx"  # Path to Excel file with names and addresses
CURR_SCRIPT_PATH = os.path.realpath(os.path.dirname(__file__))
profile_path = CURR_SCRIPT_PATH + "\\profile"  # Path to Chrome profile (you can put the full path to existing profile or keep it to create new profile and use it later)


FIRST_NAME_COL = 'F'  # (input)
LAST_NAME_COL = 'G'  # (input)
Mailing_Address_Col ='H' # (input)
Mailing_State_Col ='J'# (input)
Mailing_City_Col ='I' # (input)




PHONEs_COLs = ['M','N','O','P','Q']  # columns to output phone numbers  # (output)


######################################################################################################################


def open_chrome_with_profile():
    # Create a new Chrome session with the Chrome profile using undetected_chromedriver

    options = Options()
    options.add_argument("--user-data-dir=" + profile_path)  # Use existing profile

    # Use undetected_chromedriver to avoid detection
    driver = uc.Chrome(service=service, options=options, headless=False)  # Set headless=False for visible browser
    return driver


def open_xlsx_file():
    # Open Excel file and return the workbook and worksheet
    wb = load_workbook(filename=xlsx_path)
    ws = wb.active
    return wb, ws


def write_phones_to_xlsx_file(wb, ws, phones, row):
    # Write phones to Excel file
    print("Writing phones to Excel file...")
    for i in range(len(phones)):
        # Ensure we don't go out of range for PHONEs_COLs
        if i < len(PHONEs_COLs):
            ws[PHONEs_COLs[i] + str(row)].value = phones[i]
        else:
            # Handle cases where there are more phones than available columns
            print(f"Warning: More phones than columns. Extra phone at index {i} will not be written.")
            break  # Or use continue if you prefer to skip the rest
    wb.save(xlsx_path)
    print("Phones written to Excel file.", row)


def extract_phones_from_page(page_source, target_firstname, target_lastname):
    # Extract phones for a specific firstname and lastname from the page source
    phones = []
    try:
        # Parse the page source with BeautifulSoup
        soup = BeautifulSoup(page_source, "html.parser")
        
        # Find all div elements with the class "card" representing each person
        cards = soup.find_all("div", class_="card")
        
        for card in cards:
            # Safely extract the person's name from the card
            name_tag = card.find("span", class_="larger")
            if name_tag:
                if name_tag:
                    full_name = name_tag.get_text(strip=True).split()
                    if len(full_name) >= 2:
                        firstname = full_name[0].lower()
                        lastname = full_name[1].lower()
                        target_firstname = target_firstname.lower()
                        target_lastname = target_lastname.lower()   

                        if firstname == target_firstname and lastname == target_lastname:
                            # Find the phone numbers in this person's card
                            print("Found matching person." , firstname , lastname)
                            phone_tags = card.find_all("a", class_="nowrap")
                            for phone_tag in phone_tags:
                                phone = phone_tag.get_text(strip=True)
                                phones.append(phone)
                            break  # Stop once we find the matching person
            else:
                print("Name tag not found in card.")
        
        return phones
    
    except Exception as e:
        print(f"Error: {e}")
        return phones

def check_for_captcha_or_rate_limit(driver):
    """Check if the page contains a captcha or rate limit message"""
    page_source = driver.page_source.lower()
    captcha_indicators = [
        "captcha", 
        "rate limit", 
        "too many requests", 
        "access denied",
        "please verify you are a human",
        "please complete the security check",
        "Loading Search Results..."
    ]
    
    for indicator in captcha_indicators:
        if indicator in page_source:
            return True
    
    # Also check for common captcha elements
    try:
        captcha_elements = driver.find_elements(By.XPATH, 
            "//*[contains(@class, 'captcha') or contains(@id, 'captcha') or contains(@title, 'captcha')]")
        if captcha_elements:
            return True
    except:
        pass
    
    return False

def wait_for_user_to_solve_captcha(driver):
    """Pauses execution until the user manually solves the captcha and presses Enter"""
    print("\n" + "="*80)
    print("CAPTCHA or RATE LIMIT DETECTED!")
    print("Please solve the captcha/rate limit issue in the browser manually.")
    print("Press Enter when you've solved it to continue...")
    print("Or type 'exit' to quit the script.")
    print("="*80 + "\n")
    
    # Wait for user input
    user_input = input()
    if user_input.lower() == 'exit':
        return False
    
    print("Resuming script...")
    return True

def main():
    driver = open_chrome_with_profile()  # Open Chrome with profile using undetected_chromedriver
    driver.get("https://www.fastpeoplesearch.com/")  # Navigate to FastPeopleSearch.com

    # if access denied, wait for user to enable vpn (only for the first time)
    if "Access Denied" in driver.page_source:
        print("Access Denied")
        if not wait_for_user_to_solve_captcha(driver):
            return 1
        driver.get("https://www.fastpeoplesearch.com/")  # Navigate to FastPeopleSearch.com
        if "Access Denied" in driver.page_source:
            return 1

    wb, ws = open_xlsx_file()  # Open the Excel file
    # for each row in the Excel file search for the person and write the phones to the Excel file
    verified=False
    for row in range(2, ws.max_row + 1):
        # try searching for this person
        try:
            First_Name = ws[FIRST_NAME_COL + str(row)].value
            Last_Name = ws[LAST_NAME_COL+ str(row)].value
            Mailing_Address = ws[Mailing_Address_Col + str(row)].value
            Mailing_City = ws[Mailing_City_Col + str(row)].value
            Mailing_State = ws[Mailing_State_Col + str(row)].value
           

            if (Mailing_Address is None and Mailing_City is None and Mailing_State is None):
                continue

            # search for this person
            Mailing_Address = Mailing_Address.replace(" ", "-")
            Mailing_City = Mailing_City.replace(" ", "-")
            Mailing_State = Mailing_State.replace(" ", "-")
            driver.get("https://www.fastpeoplesearch.com/address/" + Mailing_Address + "_" + Mailing_City + "-" + Mailing_State)
            if verified==False:
                time.sleep(30)  # wait for the page to load
                verified=True
            else:
                time.sleep(1)
                
            # Check for captcha or rate limit
            if check_for_captcha_or_rate_limit(driver):
                if not wait_for_user_to_solve_captcha(driver):
                    print("Script execution terminated by user.")
                    break
            
            # try to get all phones for this person as 
            # a list of strings
            phones = extract_phones_from_page(driver.page_source , First_Name,Last_Name)
            if phones:
                # write phones to Excel file
                print("Found " + str(len(phones)) + " phones for " + Mailing_Address + " " + Mailing_City)
                write_phones_to_xlsx_file(wb, ws, phones, row)
            else:
                print("No phones found for " + Mailing_Address + " " + Mailing_City)

            # wait 1 second before searching for the next person
            time.sleep(1)

        except Exception as e:
            print(str(e))
            continue

    wb.close()
    driver.close()


if __name__ == "__main__":
    main()